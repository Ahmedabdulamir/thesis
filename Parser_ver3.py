import pandas as pd
import os
import x_section_w
import x_section
import y_intergration_w
import xlsxwriter
import openpyxl as xl
import shutil


def _opExcel(data, data_type, Path, Direction, Moment):
    
    if not os.path.exists('xlsx/' + Path + '/'+ Path + Direction + '.xlsm'):
        if not os.path.exists('xlsx/' + Path):
            os.mkdir('xlsx/' + Path)
                    
        base ='xlsx/Sheet.xlsm'
        copy ='xlsx/' + Path + '/'+ Path + Direction + '.xlsm'
        shutil.copy(base,copy)

        wb = xl.load_workbook('xlsx/' + Path + '/'+ Path + Direction + '.xlsm', read_only=False, keep_vba=True)
        ws = wb['Default']
    else:
        wb = xl.load_workbook('xlsx/' + Path + '/'+ Path + Direction + '.xlsm', read_only=False, keep_vba=True)
        ws = wb['Default']
    
    data_types = ['M', 'M_avg', 'M_h','V', 'Y']
    row_index = ['A10001', 'A10002', 'A10003','A10004','A10005']
    col_index = ['B10001', 'B10002', 'B10003','B10004','B10005']
    index = data_types.index(data_type)

    col = ws[col_index[index]].value
    row = ws[row_index[index]].value

    print(data)
    print(col)
    


    for i in range(row, row + len(data)):
        ws.cell(row=i,column=col).value = data[i-row]
        if index == 1 and i > 2:
            coor1 = _cell(ws.cell(row=i,column=col))
            coor2 = _cell(ws.cell(row=8,column=col))
            # ws.cell(row=i+8,column=col).value = '=ROUND('+ coor2 + '/' + coor1 + ',2)'
            # ws.cell(row=i+17,column=int(col/10)+140).value = '='+ _cell(ws.cell(row=i+8 ,column=col)) + '/2'
            # ws.cell(row=i+17,column=int(col/10)+155).value = '='+ _cell(ws.cell(row=i+8 ,column=col)) + '/-2'
            # ws.cell(row=19,column=int(col/10)+140).value = '='+ _cell(ws.cell(row=10 ,column=col))  
            # ws.cell(row=19,column=int(col/10)+155).value = '='+ _cell(ws.cell(row=10 ,column=col))

            ws.cell(row=i+8,column=col).value = '=ROUND('+ coor2 + '/' + coor1 + ',2)'
            ws.cell(row=int(col/10)+6,column=i+272).value = '='+ _cell(ws.cell(row=i+8 ,column=col)) + '/2'
            ws.cell(row=int(col/10)+100,column=i+272).value = '='+ _cell(ws.cell(row=i+8 ,column=col)) + '/-2'
            ws.cell(row=int(col/10)+6,column=274).value = '='+ _cell(ws.cell(row=10 ,column=col))  
            ws.cell(row=int(col/10)+100,column=274).value = '='+ _cell(ws.cell(row=10 ,column=col))  

    ws[col_index[index]].value = int(col)+2
    if index == 1 or index == 2:
       ws[col_index[index]].value = int(col)+10 
    ws['A10006'] = str(Path)
        
    wb.save('xlsx/' + Path + '/'+ Path + Direction + '.xlsm')


def _cell(coor):
    coor = str(coor)
    coor = coor.rstrip('>')
    coor = coor.split(".")
    return str(coor[1])

def _parser(path, Moment, Section, Direction, meshsize , sort, y,  w, _print):
    
    #Checking for the requested data
    if Moment == 'M2':
        Index3 =4
        if Direction == "y": 
                Index1 = 1
                Index2 = 2
        elif Direction == "x": 
            Index1 = 2
            Index2 = 1
            
    elif Moment == 'My':
        Index3 =3
        if Direction == "y": 
            Index1 = 1
            Index2 = 2
        elif Direction == "x": 
            Index1 = 2
            Index2 = 1

    elif Moment == 'Vy':
        Index3 =9
        Moment = 'My'
        if Direction == "y": 
            Index1 = 1
            Index2 = 2
        elif Direction == "x": 
            Index1 = 2
            Index2 = 1
    


    #Reading the FEM design 19 coordinates batchfile intothe node-coordinates dataframe (coor_df)
    coor_df=pd.read_csv('./temp/'+ path +'/'+ path +'_coor.txt',sep='	', header=None, engine='python')
    
    #initilizing the dataframes
    extract_load_df = pd.DataFrame()   #where the nodes corrdinates are stored   
    filtered_nodes_df = pd.DataFrame()

    node_df = coor_df.loc[abs(coor_df[Index1]-Section) <= (meshsize)]         #change the column number 1:X 2:Y      (meshsize/2)0.5= torellence
    #node_df = node_df.append(extract_df_nodes, ignore_index = True)
    
    node_df_list = node_df[0].tolist()
    node_df_list = [x for x in node_df_list if x != 'nan']

    #Reading the FEM design 19 data batchfile into the node-Load datadrame (load_df)     
    load_df=pd.read_csv('./temp/'+ path +'/'+ path + '_' + Moment + '.txt',sep='	', header=None, engine='python')


    #Extracting the related nodes data
    extract_load_df = extract_load_df.append(load_df[load_df[2].isin(node_df_list)] , ignore_index = True)
    
    #Filtering the node-coordinates dataframe according to the extracted node-load dataframe
    filtered_nodes_df = filtered_nodes_df.append(node_df[node_df[0].isin(extract_load_df[2].tolist())], ignore_index = True)
    

    #Dropping irrelvant colmuns from the node-load dataframe
    if Index3 == 9:
        dropped_col = [3, 4, 5, 6, 7, 8, 10]
        Moment = 'Vy'
    elif Index3 == 4 or Index3 == 3:
        dropped_col = list(range(6, extract_load_df.shape[1]))

    #print(extract_load_df, dropped_col)
    extract_load_df.drop(axis=1, columns=[0, 1], inplace=True)
    extract_load_df.drop(axis=1, columns=dropped_col, inplace=True)
    #print(extract_load_df, dropped_col)
    
    #Sorting and reindexing of the node-coordinates and the node-load dataframes
    filtered_nodes_df.sort_values(by=[0], inplace = True)
    extract_load_df.sort_values(by=[2], inplace = True)
    filtered_nodes_df.reset_index(drop=True, inplace=True)
    extract_load_df.reset_index(drop=True, inplace=True)

    #Merging the two dataframes and resorting according to value [sort]
    result_df = pd.concat([extract_load_df[2], extract_load_df[Index3],filtered_nodes_df[0], filtered_nodes_df[Index2], filtered_nodes_df[Index1]], axis=1)
    result_df.columns = range(result_df.shape[1])
    result_df.sort_values(by=[sort], inplace = True)
    result_df.reset_index(drop=True, inplace=True)


    # dropped_row = list(range(0 ,int((len(result_df.index))/2)))
    # result_df.drop(axis=0, index=dropped_row, inplace=True)
    # result_df.reset_index(drop=True, inplace=True)

    
    #Printing for debuging 
    if _print != False:
        print(extract_load_df, extract_load_df.shape)
        print(filtered_nodes_df)
        print(result_df)

    # Calling the interpolation function
    M_avg = [Moment + '_avg'] + [None]*len(w)
    for i in range (0, len(w)):
        W = w[i]
    # Selecting the columns to export in form of Lists as in the variables blow    
        M =  result_df[1].tolist()       #Moment/ Shear values
        Y =  result_df[3].tolist()       #Y coordinates
        No = result_df[0].tolist()      #Nodes
        X =  result_df[4].tolist()       #X coordinates
        (No, X, Y, M) = x_section_w.section_interp(Section, No, X, Y, M, W)
        _opExcel([Moment + '_X=' + str(X[1])] + list(M), 'M', path, Direction, Moment)
        _opExcel(['Y_w=' + str(W)]+ list(Y), 'Y', path, Direction, Moment)
        M_avg[i+1] = y_intergration_w.averrage (Y, M, W)
    _opExcel(M_avg, 'M_avg', path, Direction, Moment)
        
    
    # Shifting the plotting axis to the load centerline
    #Y = [x - y/2 for x in Y]

    calc = pd.read_csv('xlsx/Alpha0.5_Y0.4_X10.0/' + Moment + '.csv',  sep=',', header=None, engine='python')
    
    tolerence = 0.01
    if Moment == 'Vy':
        tolerence = 0.1
    calc = calc.loc[abs(calc[1]-Section) <= (tolerence)] 
    #print(calc)

    (No_h, X_h, Y_h, M_h) = x_section.section_interp(Section, calc[3].tolist(), calc[1].tolist(), calc[2].tolist(), calc[0].tolist())
    _opExcel(['X=' + str(X[1])] + list(M_h) + [''] + [X[1]], 'M_h', path, Direction, Moment)

    # Writing the data to the CSV file to be handled further by Excel
    #data = [[''] * 1000, [ 'L=' + str(y) + 'm', 'M_avg, w=' + str(w) + '  Sec. X=' + str(Section), M_h[1]/M_avg[1]] ,  M , Y, X, No, M_avg, M_h]
    #data = [[''] * 1000, [path], w + Y , [ 'W=' + str(y) + 'm', 'M_avg, w=' + str(w) + '  Sec. X=' + str(Section), M_h[1]/M_avg[1]] ,  M , Y, X, No, M_avg, M_h]
    #data = pd.DataFrame(data)
    #data.to_csv('temp/sections.csv', mode='a', header=False, index=False)

    return


    

def _xlsExcel(Section, Path, Direction, Moment, w, remove):
    
        
    data =pd.read_csv('temp/sections.csv',  sep=',', header=None, engine='python')
    if remove == True:
        os.remove('temp/sections.csv') 

    data.fillna('', inplace = True)
    #print(data)
    length = len(data.index)

    if not os.path.exists('xlsx/' + Path):
        os.mkdir('xlsx/' + Path)

    if os.path.isfile('xlsx/' + Path + '/w_' + str(w) + Direction + Moment + '.xlsx'):
        overwirte = input("File exists, overwirte?(Y,N) (Default = overwirte)")
        if overwirte == 'N':
            print('Please rename and close the file and retry again')
        else:
            while True:
                try:
                    os.remove('xlsx/' + Path + '/w_' + str(w) + Direction + Moment + '.xlsx')
                except IOError:
                    input('Please close Excelfile to continue')
                    continue
                break
    

    workbook = xlsxwriter.Workbook('xlsx/' + Path + '/w_' + str(w) + Direction + Moment + '.xlsx', {'strings_to_numbers': True})
    worksheet = workbook.add_worksheet()
    #chart5 = workbook.add_chart({'type': 'scatter', 'subtype': 'smooth'})
    chart2 = workbook.add_chart({'type': 'scatter', 'subtype': 'smooth'})
    chart2.set_style(15)
    
    for i in range (0 , length):
        col = xlsxwriter.utility.xl_col_to_name(i)
        pd.extdata = data.loc[(i),:]
        worksheet.write_column(col + '1', pd.extdata)
        

    for j in range (2 , length, 8):
        col0 = xlsxwriter.utility.xl_col_to_name(j-1)
        col1 = xlsxwriter.utility.xl_col_to_name(j)
        col2 = xlsxwriter.utility.xl_col_to_name(j+1)
        col3 = xlsxwriter.utility.xl_col_to_name(j+4)
        chart2.add_series({'categories': '=Sheet1!$'+ col2 +'$2:$'+ col2 +'$'+ str(len(data.columns)),
                               'values': '=Sheet1!$'+ col1 +'$2:$'+ col1 +'$'+ str(len(data.columns)),
                                 'name': '=Sheet1!$'+ col0 +'$1'})
        chart2.add_series({'categories': '=Sheet1!$'+ col2 +'$2:$'+ col2 +'$'+ str(len(data.columns)),
                               'values': '=Sheet1!$'+ col3 +'$2:$'+ col3 +'$'+ str(len(data.columns)),
                                 'name': '=Sheet1!$'+ col0 +'$1'})
       

    # Add the worksheet data to be plotted.
    
    # Create a new chart object.
   

    # Add a series to the chart.
   

    # Insert the chart into the worksheet.
    worksheet.insert_chart('H1', chart2)

    workbook.close()
    
